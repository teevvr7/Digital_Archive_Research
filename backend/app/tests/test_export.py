"""Tests for the export module — CSV/XLSX row export and zip bulk-download.

All I/O is mocked — no DB, no Supabase, matching the rest of this project's
test style.
"""

import csv
import datetime
import io
import uuid
import zipfile
from unittest.mock import MagicMock, patch

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.models.document import Document
from app.modules.export import service


def _make_doc(
    *,
    title: str = "Invoice ABC.pdf",
    original_filename: str = "invoice.pdf",
    vendor: str | None = "Acme Corp",
    invoice_no: str | None = "INV-1",
    total_amount: float | None = 100.0,
    currency: str | None = "MYR",
    storage_key: str = "tenant/docs/invoice.pdf",
) -> Document:
    doc = MagicMock(spec=Document)
    doc.id = uuid.uuid4()
    doc.title = title
    doc.original_filename = original_filename
    doc.vendor = vendor
    doc.invoice_no = invoice_no
    doc.total_amount = total_amount
    doc.currency = currency
    doc.document_type = "invoice"
    doc.status = "completed"
    doc.document_date = datetime.date(2026, 7, 1)
    doc.uploaded_at = datetime.datetime(2026, 7, 1, 9, 0, tzinfo=datetime.timezone.utc)
    doc.storage_key = storage_key
    doc.deleted_at = None
    return doc


# ---------------------------------------------------------------------------
# export_documents
# ---------------------------------------------------------------------------

class TestExportDocuments:
    def test_invalid_format_raises_400(self) -> None:
        db = MagicMock()
        with pytest.raises(HTTPException) as exc_info:
            service.export_documents(db, fmt="pdf")
        assert exc_info.value.status_code == 400

    def test_csv_export_contains_expected_rows(self) -> None:
        db = MagicMock()
        docs = [_make_doc(), _make_doc(title="Receipt XYZ.pdf", vendor="Beta Sdn Bhd", invoice_no="RCP-2", total_amount=50.5)]
        db.scalars.return_value.all.return_value = docs

        content, media_type, filename, truncated = service.export_documents(db, fmt="csv")

        assert media_type == "text/csv"
        assert filename.startswith("documents-") and filename.endswith(".csv")
        assert truncated is False

        text = content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[0] == service._COLUMNS
        assert rows[1][0] == "Invoice ABC.pdf"
        assert rows[1][1] == "Acme Corp"
        assert rows[1][2] == "INV-1"
        assert rows[1][3] == "100.00"
        assert rows[2][1] == "Beta Sdn Bhd"

    def test_csv_handles_missing_typed_fields_gracefully(self) -> None:
        db = MagicMock()
        doc = _make_doc(vendor=None, invoice_no=None, total_amount=None, currency=None)
        db.scalars.return_value.all.return_value = [doc]

        content, _, _, _ = service.export_documents(db, fmt="csv")
        rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        assert rows[1][1] == ""  # vendor
        assert rows[1][3] == ""  # total amount

    def test_xlsx_export_opens_and_has_correct_rows(self) -> None:
        db = MagicMock()
        docs = [_make_doc()]
        db.scalars.return_value.all.return_value = docs

        content, media_type, filename, truncated = service.export_documents(db, fmt="xlsx")

        assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert filename.endswith(".xlsx")
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == tuple(service._COLUMNS)
        assert rows[1][0] == "Invoice ABC.pdf"
        assert rows[1][2] == "INV-1"

    def test_truncation_flag_set_when_over_limit(self) -> None:
        db = MagicMock()
        # One more row than the limit — export_documents should detect and slice it off.
        docs = [_make_doc() for _ in range(service._EXPORT_ROW_LIMIT + 1)]
        db.scalars.return_value.all.return_value = docs

        content, _, _, truncated = service.export_documents(db, fmt="csv")
        assert truncated is True
        rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        assert len(rows) - 1 == service._EXPORT_ROW_LIMIT  # minus header

    def test_no_truncation_at_exact_limit(self) -> None:
        db = MagicMock()
        docs = [_make_doc() for _ in range(service._EXPORT_ROW_LIMIT)]
        db.scalars.return_value.all.return_value = docs

        _, _, _, truncated = service.export_documents(db, fmt="csv")
        assert truncated is False


# ---------------------------------------------------------------------------
# bulk_download_zip
# ---------------------------------------------------------------------------

class TestBulkDownloadZip:
    def test_empty_list_raises_400(self) -> None:
        db = MagicMock()
        with pytest.raises(HTTPException) as exc_info:
            service.bulk_download_zip(db, [])
        assert exc_info.value.status_code == 400

    def test_over_limit_raises_413(self) -> None:
        db = MagicMock()
        ids = [uuid.uuid4() for _ in range(service._MAX_BULK_DOWNLOAD + 1)]
        with pytest.raises(HTTPException) as exc_info:
            service.bulk_download_zip(db, ids)
        assert exc_info.value.status_code == 413

    def test_builds_zip_with_correct_entries(self) -> None:
        db = MagicMock()
        doc1 = _make_doc(original_filename="a.pdf", storage_key="k1")
        doc2 = _make_doc(original_filename="b.pdf", storage_key="k2")
        db.scalars.return_value.all.return_value = [doc1, doc2]

        with patch(
            "app.modules.export.service.object_storage.download_file",
            side_effect=lambda key: b"content-" + key.encode(),
        ):
            content = service.bulk_download_zip(db, [doc1.id, doc2.id])

        zf = zipfile.ZipFile(io.BytesIO(content))
        names = zf.namelist()
        assert "a.pdf" in names
        assert "b.pdf" in names
        assert zf.read("a.pdf") == b"content-k1"
        assert zf.read("b.pdf") == b"content-k2"

    def test_duplicate_filenames_get_disambiguated(self) -> None:
        db = MagicMock()
        doc1 = _make_doc(original_filename="invoice.pdf", storage_key="k1")
        doc2 = _make_doc(original_filename="invoice.pdf", storage_key="k2")
        db.scalars.return_value.all.return_value = [doc1, doc2]

        with patch(
            "app.modules.export.service.object_storage.download_file",
            return_value=b"x",
        ):
            content = service.bulk_download_zip(db, [doc1.id, doc2.id])

        names = zipfile.ZipFile(io.BytesIO(content)).namelist()
        assert len(names) == 2
        assert len(set(names)) == 2  # no collision

    def test_skips_file_that_fails_to_download(self) -> None:
        db = MagicMock()
        doc1 = _make_doc(original_filename="ok.pdf", storage_key="k1")
        doc2 = _make_doc(original_filename="broken.pdf", storage_key="k2")
        db.scalars.return_value.all.return_value = [doc1, doc2]

        def _download(key: str) -> bytes:
            if key == "k2":
                raise RuntimeError("storage unavailable")
            return b"fine"

        with patch(
            "app.modules.export.service.object_storage.download_file",
            side_effect=_download,
        ):
            content = service.bulk_download_zip(db, [doc1.id, doc2.id])

        names = zipfile.ZipFile(io.BytesIO(content)).namelist()
        assert "ok.pdf" in names
        assert "broken.pdf" not in names
